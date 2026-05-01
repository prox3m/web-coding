from django.test import TestCase, RequestFactory
from django.contrib.admin.sites import AdminSite
from django.contrib.auth.models import User
from game.admin import export_to_xlsx
from game.models import GameSession
from io import BytesIO
import openpyxl

class AdminXLSXExportTest(TestCase):
    def setUp(self):
        self.site = AdminSite()
        self.factory = RequestFactory()
        self.admin_user = User.objects.create_user(username='xlsx_admin', password='pass', is_staff=True, is_superuser=True)
        self.session = GameSession.objects.create(
            user=self.admin_user, game_state={}, score=750, level=4, time_played=180, is_completed=True
        )

    def test_xlsx_export_structure_and_content(self):
        queryset = GameSession.objects.filter(id=self.session.id)
        request = self.factory.get('/admin/game/gamesession/')
        request.user = self.admin_user

        response = export_to_xlsx(None, request, queryset)
        
        # Проверка заголовков
        self.assertEqual(response['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        self.assertIn('attachment; filename=arkanoid_sessions_', response['Content-Disposition'])

        # Парсинг файла
        wb = openpyxl.load_workbook(filename=BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws['A1'].value, 'ID')
        self.assertEqual(ws['B1'].value, 'Игрок')
        self.assertEqual(ws['C1'].value, 'Очки')
        
        # Проверка данных
        self.assertEqual(ws['A2'].value, self.session.id)
        self.assertEqual(ws['B2'].value, self.session.user.username)
        self.assertEqual(ws['C2'].value, 750)
        self.assertEqual(ws['F2'].value, 'Да')  # is_completed